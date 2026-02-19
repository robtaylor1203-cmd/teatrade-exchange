"""
Mombasa Auction Data Importer
==============================
Reads the two Excel formats from the test_data folder and loads them into
Supabase.  Run this script whenever a new auction report is received.

Usage:
    python import_auction_data.py --general  "GeneralReport (87).xlsx"
    python import_auction_data.py --quantity "Auction Quantity 2025.xlsx"
    python import_auction_data.py --general  "GeneralReport (87).xlsx" \
                                  --quantity "Auction Quantity 2025.xlsx"

Requirements:
    pip install openpyxl requests python-dotenv

Configuration:
    Either set SUPABASE_URL and SUPABASE_SERVICE_KEY as environment variables,
    or create a .env file in the project root containing those two keys.
    The SERVICE KEY (not the anon key) is required so the script can bypass
    Row-Level Security and write directly to the auction tables.
"""

import argparse
import os
import sys
import json
import re
from datetime import datetime, date
from collections import defaultdict
from pathlib import Path

import openpyxl
import requests

try:
    from dotenv import load_dotenv
    load_dotenv(Path(__file__).parent.parent / ".env")
except ImportError:
    pass  # dotenv is optional

# ─────────────────────────────────────────────────────────────────────────────
# Configuration
# ─────────────────────────────────────────────────────────────────────────────
SUPABASE_URL = os.getenv("SUPABASE_URL", "https://uznxzyuknigzlxecjgtb.supabase.co")
SERVICE_KEY  = os.getenv("SUPABASE_SERVICE_KEY", "")  # must be the service_role key

HEADERS = {
    "apikey":        SERVICE_KEY,
    "Authorization": f"Bearer {SERVICE_KEY}",
    "Content-Type":  "application/json",
    "Prefer":        "return=representation",
}

# Grade → tea symbol mapping.
# Each entry means: compute a per-grade VWAP and store it in
# auction_sales.<col>_vwap, then use it as the anchor for that symbol.
GRADE_TO_TEA_SYMBOL = {
    "PF1":   ("KEN-PF1",  "pf1_vwap"),
    "BP1":   ("KEN-BP1",  "bp1_vwap"),
    "DUST1": ("KEN-DUST", "dust1_vwap"),
    "PD":    ("KEN-PD",   "pd_vwap"),
    "BMF":   ("KEN-BMF",  "bmf_vwap"),
    # FNGS1 is the higher-quality fannings; use it as the KEN-FNGS anchor.
    # Also accept plain FNGS as a fallback.
    "FNGS1": ("KEN-FNGS", "fngs_vwap"),
    "FNGS":  ("KEN-FNGS", "fngs_vwap"),
}

# All Kenya grades used for the KENYA / MOMBASA composite index VWAP
KENYA_GRADES = {"BP1", "PF1", "DUST1", "PD", "DUST", "PF", "BP",
                "BMF", "BMF1", "DUST2", "FNGS", "FNGS1", "PF2"}


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────
def _api(method: str, table: str, payload=None, params: str = ""):
    url = f"{SUPABASE_URL}/rest/v1/{table}{params}"
    resp = requests.request(method, url, headers=HEADERS,
                            json=payload, timeout=30)
    if not resp.ok:
        print(f"  [ERROR] {method} {table}: {resp.status_code} {resp.text[:300]}")
        return None
    return resp.json() if resp.text else {}


def _insert(table: str, rows: list, chunk: int = 500):
    """Insert rows in chunks; return list of inserted records."""
    results = []
    for i in range(0, len(rows), chunk):
        batch = rows[i:i + chunk]
        r = _api("POST", table, batch)
        if r:
            results.extend(r if isinstance(r, list) else [r])
    return results


def _upsert(table: str, rows: list, on_conflict: str, chunk: int = 500):
    """Upsert rows — PostgREST requires conflict columns in the URL query string."""
    results = []
    for i in range(0, len(rows), chunk):
        batch = rows[i:i + chunk]
        url = f"{SUPABASE_URL}/rest/v1/{table}?on_conflict={on_conflict}"
        h = {**HEADERS, "Prefer": "resolution=merge-duplicates,return=representation"}
        resp = requests.post(url, headers=h, json=batch, timeout=30)
        if resp.ok and resp.text:
            r = resp.json()
            results.extend(r if isinstance(r, list) else [r])
        elif not resp.ok:
            print(f"  [ERROR] upsert {table}: {resp.status_code} {resp.text[:300]}")
    return results


def _call_rpc(fn: str, payload: dict):
    """Call a Supabase RPC (PostgreSQL function)."""
    return _api("POST", f"rpc/{fn}", payload, "")


def vwap(prices_weights: list) -> float | None:
    """Compute volume-weighted average price.  Each item is (price, weight)."""
    total_w = sum(w for _, w in prices_weights if w)
    if total_w == 0:
        return None
    return sum(p * w for p, w in prices_weights if w) / total_w


def num(value):
    """Coerce a value to float, returning None for empty strings / non-numeric."""
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def intnum(value):
    """Coerce a value to int, returning None for empty/non-numeric (for INTEGER columns)."""
    v = num(value)
    return int(v) if v is not None else None


def txt(value):
    """Coerce a value to a non-empty string, returning None for blank cells."""
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


# ─────────────────────────────────────────────────────────────────────────────
# GeneralReport importer
# ─────────────────────────────────────────────────────────────────────────────
def parse_selling_end_time(raw) -> str | None:
    """Try to parse the Selling End Time field into an ISO timestamp."""
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw.isoformat()
    s = str(raw).strip()
    # Format observed: "16/09/2025 08:15:32:963"  (last group is milliseconds)
    m = re.match(r"(\d{2}/\d{2}/\d{4}) (\d{2}:\d{2}:\d{2}):\d+", s)
    if m:
        try:
            return datetime.strptime(f"{m.group(1)} {m.group(2)}",
                                     "%d/%m/%Y %H:%M:%S").isoformat()
        except ValueError:
            pass
    return s if s else None  # return None for blank cells


def import_general_report(filepath: str):
    """
    Parse a GeneralReport Excel file and load:
      • auction_sales          (1 row)
      • auction_lots           (1 row per lot)
      • auction_grade_summary  (1 row per grade)
    Then call apply_auction_prices_to_teas() to update price history.
    """
    print(f"\n{'='*60}")
    print(f"Importing GeneralReport: {filepath}")
    print(f"{'='*60}")

    if not SERVICE_KEY:
        print("[ERROR] SUPABASE_SERVICE_KEY is not set.  "
              "Export it as an env var before running this script.")
        sys.exit(1)

    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    ws = wb["General Report"]

    # ── 1. Collect all rows ──────────────────────────────────────────────────
    header_row1 = None  # col names
    header_row2 = None  # sub-names (row 2)
    lots_raw = []

    for i, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
        if i == 1:
            header_row1 = list(row)
        elif i == 2:
            header_row2 = list(row)
        else:
            if row[0] is None:
                continue
            lots_raw.append(row)

    print(f"  Rows read: {len(lots_raw)}")

    # Column indices (0-based) from the header we observed:
    # 0:Broker 1:LotNo 2:SellingMark 3:Grade 4:InvoiceNo 5:SubElev
    # 6:SaleCode 7:Category 8:RP 9:RA 10:Certifications
    # 11:Bags 12:NetWeight 13:TotalWeight
    # 14-23: Quality adjectives / standards
    # 25:AskingPrice 26:BaselinePrice
    # 27:RegisteredBidAmount 28:RegBidBuyerCode 29:RegBidBuyerCompany
    # 30:SecondHighestBid 31:2ndBuyerCode 32:2ndBuyerCompany
    # 33:TotalPrice 34:Status 35:PurchasedPrice
    # 36:BuyerCode 37:BuyerName
    # 38:Factory 39:ProducerCountry 40:WarehouseCompany 41:WarehouseLocation
    # 43:ManufacturedDate 44:OutlotType 45:SellingEndTime 46:Producer
    # 47:FinalBuyerName 48:FinalBuyerCode 49:FinalBuyerUser
    # 50:FinalPrice 51:TotalValue 52:TransactionType
    C = {
        "broker": 0, "lot_no": 1, "selling_mark": 2, "grade": 3,
        "invoice_no": 4, "sub_elevation": 5, "sale_code": 6, "category": 7,
        "rp": 8, "ra": 9, "certifications": 10,
        "bags": 11, "net_weight": 12, "total_weight": 13,
        "asking_price": 25, "baseline_price": 26,
        "reg_bid_amount": 27, "reg_bid_buyer": 28,
        "second_bid": 30, "second_buyer": 31,
        "total_price": 33, "status": 34, "purchased_price": 35,
        "buyer_code": 36, "buyer_name": 37,
        "factory": 38, "producer_country": 39,
        "warehouse_company": 40, "warehouse_location": 41,
        "manufactured_date": 43, "selling_end_time": 44,
        "producer": 45, "final_buyer_name": 46,
        "final_price": 49, "total_value": 50, "transaction_type": 51,
    }

    def g(row, key):
        idx = C.get(key)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    # ── 2. Determine sale identity from first valid row ──────────────────────
    sale_code_full = g(lots_raw[0], "sale_code")  # "Sale 37 - M2"
    sale_code = re.match(r"(Sale \d+)", sale_code_full or "").group(1) \
        if sale_code_full else "Unknown"
    sale_number = int(re.search(r"\d+", sale_code).group()) \
        if re.search(r"\d+", sale_code) else 0

    # Best-effort sale date from the earliest selling_end_time
    sale_date = None
    for row in lots_raw:
        raw_ts = g(row, "selling_end_time")
        parsed = parse_selling_end_time(raw_ts)
        if parsed:
            try:
                sale_date = datetime.fromisoformat(parsed).date().isoformat()
                break
            except (ValueError, TypeError):
                pass

    print(f"  Sale: {sale_code}  |  Date: {sale_date}")

    # ── 3. Aggregate stats ───────────────────────────────────────────────────
    status_counts = defaultdict(int)
    grade_data = defaultdict(lambda: {
        "lots_sold": 0, "lots_unsold": 0,
        "weight_price_pairs": [],   # (price, weight) for VWAP
        "prices": [],
        "total_value": 0.0,
        "total_weight_sold": 0.0,
    })

    lot_records = []

    for row in lots_raw:
        status  = g(row, "status") or ""
        grade   = (g(row, "grade") or "").strip()
        country = g(row, "producer_country") or ""
        is_sold = status in ("Sold", "Private Sold")
        price   = g(row, "purchased_price")
        weight  = g(row, "total_weight")
        value   = g(row, "total_value")

        status_counts[status] += 1

        if is_sold and price and weight and grade:
            gd = grade_data[grade]
            gd["lots_sold"] += 1
            gd["weight_price_pairs"].append((float(price), float(weight)))
            gd["prices"].append(float(price))
            gd["total_value"] += float(value or 0)
            gd["total_weight_sold"] += float(weight)
        elif grade:
            grade_data[grade]["lots_unsold"] += 1

        # Build lot record — use num() for every numeric column and txt() for text
        # to ensure empty Excel cells become NULL rather than "" which Postgres rejects.
        lot_records.append({
            "sale_number":           sale_number,
            "broker_code":           txt(g(row, "broker")),
            "lot_number":            intnum(g(row, "lot_no")),
            "selling_mark":          txt(g(row, "selling_mark")),
            "grade":                 grade,
            "invoice_no":            txt(g(row, "invoice_no")),
            "sub_elevation":         txt(g(row, "sub_elevation")),
            "category":              txt(g(row, "category")),
            "rp":                    txt(g(row, "rp")),
            "ra":                    txt(g(row, "ra")),
            "certifications":        txt(g(row, "certifications")),
            "bags":                  intnum(g(row, "bags")),
            "net_weight_per_bag_kg": num(g(row, "net_weight")),
            "total_weight_kg":       num(g(row, "total_weight")),
            "asking_price":          num(g(row, "asking_price")),
            "baseline_price":        num(g(row, "baseline_price")),
            "registered_bid_price":  num(g(row, "reg_bid_amount")),
            "registered_bid_buyer":  txt(g(row, "reg_bid_buyer")),
            "second_highest_bid":    num(g(row, "second_bid")),
            "second_highest_buyer":  txt(g(row, "second_buyer")),
            "total_price":           num(g(row, "total_price")),
            "status":                txt(status),
            "purchased_price":       num(price),
            "buyer_code":            txt(g(row, "buyer_code")),
            "buyer_name":            txt(g(row, "buyer_name")),
            "factory":               txt(g(row, "factory")),
            "producer_country":      txt(country),
            "warehouse_company":     txt(g(row, "warehouse_company")),
            "warehouse_location":    txt(g(row, "warehouse_location")),
            "manufactured_date":     txt(g(row, "manufactured_date")),
            "selling_end_time":      parse_selling_end_time(g(row, "selling_end_time")),
            "producer":              txt(g(row, "producer")),
            "final_buyer_name":      txt(g(row, "final_buyer_name")),
            "final_price":           num(g(row, "final_price")),
            "total_value":           num(value),
            "transaction_type":      txt(g(row, "transaction_type")),
        })

    # ── 4. Compute sale-level aggregates ─────────────────────────────────────
    all_pw = [(p, w) for gd in grade_data.values() for p, w in gd["weight_price_pairs"]]
    kenya_pw = [
        (p, w) for g_name, gd in grade_data.items()
        if g_name in KENYA_GRADES
        for p, w in gd["weight_price_pairs"]
    ]

    total_weight_sold = sum(w for _, w in all_pw)
    total_value_sold  = sum(gd["total_value"] for gd in grade_data.values())

    sale_vwap         = vwap(all_pw)
    kenya_index_price = vwap(kenya_pw)

    # Compute a VWAP for every grade that maps to a tea symbol.
    # FNGS1 takes priority over FNGS for the KEN-FNGS column.
    grade_vwaps: dict[str, float | None] = {}
    seen_col: set[str] = set()
    for grade, (sym, col) in GRADE_TO_TEA_SYMBOL.items():
        if col in seen_col:          # e.g. FNGS already set by FNGS1
            continue
        v = vwap(grade_data[grade]["weight_price_pairs"])
        if v is not None:
            grade_vwaps[col] = v
            seen_col.add(col)

    def r4(v): return round(v, 4) if v is not None else None

    print(f"  Lots: {len(lot_records)} total | "
          f"{status_counts['Sold']} sold | {status_counts['Unsold']} unsold")
    print(f"  Total weight sold: {total_weight_sold:,.0f} kg")
    print(f"  VWAP all grades:   USD {sale_vwap:.4f}/kg")
    print(f"  Kenya index:       USD {kenya_index_price:.4f}/kg")
    for col, v in sorted(grade_vwaps.items()):
        print(f"  {col:<12}: USD {v:.4f}/kg")

    # ── 5. Upsert auction_sales ──────────────────────────────────────────────
    sale_payload = {
        "sale_number":       sale_number,
        "sale_code":         sale_code,
        "sale_date":         sale_date,
        "total_lots":        len(lot_records),
        "lots_sold":         status_counts["Sold"] + status_counts["Private Sold"],
        "lots_unsold":       status_counts["Unsold"],
        "lots_outsold":      status_counts["Outsold"],
        "lots_private":      status_counts["Private Sold"],
        "total_weight_kg":   round(total_weight_sold, 2),
        "total_value":       round(total_value_sold, 4),
        "vwap_usd_per_kg":   r4(sale_vwap),
        "kenya_index_price": r4(kenya_index_price),
        # Per-grade VWAP columns (all the grades we track)
        **{col: r4(v) for col, v in grade_vwaps.items()},
    }

    print("\n  Upserting auction_sales…")
    inserted_sales = _upsert("auction_sales", [sale_payload], "sale_number")
    if not inserted_sales:
        print("  [ERROR] Failed to upsert sale record. Aborting.")
        return

    sale_id = inserted_sales[0]["id"]
    print(f"  sale_id = {sale_id}")

    # ── 6. Delete existing lots for this sale (idempotent re-import) ─────────
    print("  Removing any existing lots for this sale…")
    _api("DELETE", "auction_lots",
         params=f"?sale_number=eq.{sale_number}")

    # ── 7. Attach sale_id to lots and insert ──────────────────────────────────
    for lot in lot_records:
        lot["sale_id"] = sale_id

    print(f"  Inserting {len(lot_records)} auction_lots…")
    _insert("auction_lots", lot_records)

    # ── 8. Build and upsert grade summaries ──────────────────────────────────
    summary_rows = []
    for grade, gd in grade_data.items():
        if not gd["prices"]:
            continue
        gv = vwap(gd["weight_price_pairs"])
        summary_rows.append({
            "sale_id":         sale_id,
            "sale_number":     sale_number,
            "sale_date":       sale_date,
            "grade":           grade,
            "lots_sold":       gd["lots_sold"],
            "lots_unsold":     gd["lots_unsold"],
            "total_weight_kg": round(gd["total_weight_sold"], 2),
            "vwap":            round(gv, 4) if gv else None,
            "avg_price":       round(sum(gd["prices"]) / len(gd["prices"]), 4),
            "min_price":       min(gd["prices"]),
            "max_price":       max(gd["prices"]),
            "total_value":     round(gd["total_value"], 4),
        })

    print(f"  Upserting {len(summary_rows)} grade summaries…")
    _upsert("auction_grade_summary", summary_rows, "sale_id,grade")

    # ── 9. Apply prices to teas table + price_history ────────────────────────
    print("  Applying auction prices to teas and price_history…")
    _call_rpc("apply_auction_prices_to_teas", {"p_sale_id": sale_id})

    print(f"\n  [OK] Import complete for {sale_code}")


# ─────────────────────────────────────────────────────────────────────────────
# Auction Quantity importer
# ─────────────────────────────────────────────────────────────────────────────
def import_auction_quantity(filepath: str):
    """
    Parse the Auction Quantity Excel file.
    Layout: rows = weekly sales, columns grouped by year.
    Groups of 8 columns per year: [Total, Main, Sec, Kenya, Foreign, Reprints, Fresh, (blank)]
    Header row 1 has year labels like "Ttl'25", "Ttl'24" etc.
    Column 0 is "From Date/Sale No" in format "06.01.25 sale 01"
    """
    print(f"\n{'='*60}")
    print(f"Importing Auction Quantity: {filepath}")
    print(f"{'='*60}")

    if not SERVICE_KEY:
        print("[ERROR] SUPABASE_SERVICE_KEY is not set.")
        sys.exit(1)

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb["Sheet1"]

    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    # Identify year-column groups: each group starts with "Ttl'YY"
    year_groups = []  # list of (year, col_offset_0based) where col_offset points to Total
    for i, h in enumerate(headers):
        # Match many observed formats: "Ttl'25", "Ttl '16", "Ttl. '14", "  Ttl.·13"
        if h and isinstance(h, str) and re.search(r"[Tt]tl", h):
            yy_match = re.search(r"(\d{2})$", h.strip())
            if yy_match:
                yy = int(yy_match.group(1))
                year = 2000 + yy
                year_groups.append((year, i))  # i is 0-based column of Total

    print(f"  Years found: {[y for y, _ in year_groups]}")

    rows_to_insert = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        date_label = row[0]
        if not date_label:
            continue

        # Parse date and sale number: "06.01.25 sale 01"
        m = re.match(r"(\d{2})\.(\d{2})\.(\d{2})\s+sale\s+(\d+)", str(date_label), re.I)
        if not m:
            continue
        day, month, yy, sale_no = int(m.group(1)), int(m.group(2)), \
                                   int(m.group(3)), int(m.group(4))
        full_year = 2000 + yy
        sale_date = date(full_year, month, day).isoformat()

        for year, col in year_groups:
            # Columns: Total, Main, Sec, Kenya, Foreign, Reprints, Fresh
            def safe(offset):
                idx = col + offset
                return row[idx] if idx < len(row) else None

            total     = safe(0)
            main      = safe(1)
            secondary = safe(2)
            kenya     = safe(3)
            foreign   = safe(4)
            reprints  = safe(5)
            fresh     = safe(6)

            # Skip entirely empty year columns for this row
            if all(v is None for v in [total, main, secondary, kenya, foreign]):
                continue

            rows_to_insert.append({
                "sale_date":      sale_date,
                "sale_number":    sale_no,
                "year":           year,
                "total_bags":     total,
                "main_bags":      main,
                "secondary_bags": secondary,
                "kenya_bags":     kenya,
                "foreign_bags":   foreign,
                "reprints_bags":  reprints,
                "fresh_bags":     fresh,
            })

    print(f"  Rows to insert: {len(rows_to_insert)}")
    _upsert("auction_weekly_volumes", rows_to_insert, "sale_date,year")
    print("  [OK] Import complete for Auction Quantity")


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description="Import Mombasa auction data into Supabase.\n"
                    "When no flags are given, ALL GeneralReport*.xlsx files\n"
                    "in the script folder are imported oldest-first so the\n"
                    "database accumulates the full history without overwriting.")
    parser.add_argument("--general",  metavar="FILE", action="append",
                        help="Path to a GeneralReport Excel file (repeatable)")
    parser.add_argument("--quantity", metavar="FILE",
                        help="Path to an Auction Quantity Excel file")
    args = parser.parse_args()

    script_dir = Path(__file__).parent

    # ── Auction Quantity (volumes) ───────────────────────────────────────────
    quantity_file = args.quantity
    if not quantity_file:
        candidates_q = sorted(script_dir.glob("Auction Quantity*.xlsx"))
        if candidates_q:
            quantity_file = str(candidates_q[-1])
            print(f"Auto-detected Auction Quantity: {quantity_file}")

    if quantity_file:
        import_auction_quantity(quantity_file)

    # ── GeneralReport files (lots + prices) ──────────────────────────────────
    # Accept explicit list or auto-detect ALL files in the folder.
    # Sort alphabetically — file names typically include the sale number
    # (e.g. "GeneralReport (37).xlsx") so alphabetical = chronological.
    general_files = args.general or []
    if not general_files:
        general_files = sorted(str(p) for p in script_dir.glob("GeneralReport*.xlsx"))
        if general_files:
            print(f"Auto-detected {len(general_files)} GeneralReport file(s):")
            for f in general_files:
                print(f"  {Path(f).name}")

    if not general_files and not quantity_file:
        parser.print_help()
        sys.exit(1)

    for filepath in general_files:
        import_general_report(filepath)


if __name__ == "__main__":
    main()
