"""Quick dry-run to verify the import script parses the Excel files correctly
without actually talking to Supabase.  Safe to run any time."""
import sys, re
from pathlib import Path
from datetime import datetime, date
from collections import defaultdict
import openpyxl

BASE = Path(__file__).parent

KENYA_GRADES = {"BP1", "PF1", "DUST1", "PD", "DUST", "PF", "BP",
                "BMF", "BMF1", "DUST2", "FNGS", "FNGS1", "PF2"}

def vwap(pairs):
    tw = sum(w for _, w in pairs if w)
    return sum(p*w for p,w in pairs if w)/tw if tw else None

def parse_selling_end_time(raw):
    if raw is None: return None
    if isinstance(raw, datetime): return raw.isoformat()
    s = str(raw).strip()
    m = re.match(r"(\d{2}/\d{2}/\d{4}) (\d{2}:\d{2}:\d{2}):\d+", s)
    if m:
        try:
            return datetime.strptime(f"{m.group(1)} {m.group(2)}", "%d/%m/%Y %H:%M:%S").isoformat()
        except ValueError: pass
    return s

# ── GeneralReport ──────────────────────────────────────────────────────────
gr_file = BASE / "GeneralReport (87).xlsx"
print(f"Parsing {gr_file.name}…")
wb = openpyxl.load_workbook(gr_file, data_only=True, read_only=True)
ws = wb["General Report"]
C = {"broker":0,"lot_no":1,"selling_mark":2,"grade":3,"invoice_no":4,
     "sub_elevation":5,"sale_code":6,"category":7,"rp":8,"ra":9,
     "certifications":10,"bags":11,"net_weight":12,"total_weight":13,
     "asking_price":25,"baseline_price":26,"reg_bid_amount":27,"reg_bid_buyer":28,
     "second_bid":30,"second_buyer":31,"total_price":33,"status":34,
     "purchased_price":35,"buyer_code":36,"buyer_name":37,"factory":38,
     "producer_country":39,"warehouse_company":40,"warehouse_location":41,
     "manufactured_date":43,"selling_end_time":44,"producer":45,
     "final_buyer_name":46,"final_price":49,"total_value":50,"transaction_type":51}
def g(row, key): idx=C.get(key); return row[idx] if idx is not None and idx<len(row) else None

lots_raw = [r for i,r in enumerate(ws.iter_rows(min_row=3, values_only=True)) if r[0]]
sale_code_full = g(lots_raw[0], "sale_code")
sale_code = re.match(r"(Sale \d+)", sale_code_full or "").group(1)
sale_number = int(re.search(r"\d+", sale_code).group())

sale_date = None
for row in lots_raw:
    parsed = parse_selling_end_time(g(row, "selling_end_time"))
    if parsed:
        try: sale_date = datetime.fromisoformat(parsed).date().isoformat(); break
        except: pass

status_counts = defaultdict(int)
grade_data = defaultdict(lambda: {"lots_sold":0,"lots_unsold":0,"pairs":[],"prices":[],"val":0.0,"weight":0.0})

for row in lots_raw:
    status = g(row,"status") or ""
    grade  = (g(row,"grade") or "").strip()
    price  = g(row,"purchased_price")
    weight = g(row,"total_weight")
    value  = g(row,"total_value")
    status_counts[status] += 1
    if status in ("Sold","Private Sold") and price and weight and grade:
        gd = grade_data[grade]
        gd["lots_sold"] += 1
        gd["pairs"].append((float(price), float(weight)))
        gd["prices"].append(float(price))
        gd["val"] += float(value or 0)
        gd["weight"] += float(weight)
    elif grade:
        grade_data[grade]["lots_unsold"] += 1

all_pw    = [(p,w) for gd in grade_data.values() for p,w in gd["pairs"]]
kenya_pw  = [(p,w) for gn,gd in grade_data.items() if gn in KENYA_GRADES for p,w in gd["pairs"]]

tw  = sum(w for _,w in all_pw)
tv  = sum(gd["val"] for gd in grade_data.values())
overall_vwap      = vwap(all_pw)
kenya_idx         = vwap(kenya_pw)
pf1_vwap  = vwap(grade_data["PF1"]["pairs"])
bp1_vwap  = vwap(grade_data["BP1"]["pairs"])
dust1_vwap = vwap(grade_data["DUST1"]["pairs"])

print(f"\n{'='*56}")
print(f"  {sale_code}   |   Date: {sale_date}")
print(f"{'='*56}")
print(f"  Total lots:         {len(lots_raw):>8,}")
print(f"  Sold:               {status_counts['Sold']:>8,}")
print(f"  Private Sold:       {status_counts['Private Sold']:>8,}")
print(f"  Unsold:             {status_counts['Unsold']:>8,}")
print(f"  Outsold:            {status_counts['Outsold']:>8,}")
print(f"  Total weight sold:  {tw:>12,.0f} kg")
print(f"  Total value:        {tv:>14,.2f} USD")
print()
print(f"  VWAP all grades:    USD {overall_vwap:.4f}/kg")
print(f"  Kenya index price:  USD {kenya_idx:.4f}/kg")
print()
print(f"  {'Grade':<8} {'Lots':>5} {'Weight KG':>12} {'VWAP':>9} {'Min':>7} {'Max':>7}")
print(f"  {'-'*52}")
for grade, gd in sorted(grade_data.items()):
    if not gd["prices"]: continue
    gv = vwap(gd["pairs"])
    print(f"  {grade:<8} {gd['lots_sold']:>5} {gd['weight']:>12,.0f} "
          f"{gv:>9.4f} {min(gd['prices']):>7.3f} {max(gd['prices']):>7.3f}")

print()
print(f"  Tea symbol anchors after import:")
print(f"    KEN-PF1  = USD {pf1_vwap:.4f}/kg")
print(f"    KEN-BP1  = USD {bp1_vwap:.4f}/kg")
print(f"    KEN-DUST = USD {dust1_vwap:.4f}/kg")
print(f"    KENYA/MOMBASA index = USD {kenya_idx:.4f}/kg")

# ── Auction Quantity ────────────────────────────────────────────────────────
aq_file = BASE / "Auction Quantity 2025.xlsx"
print(f"\n\nParsing {aq_file.name}…")
wb2 = openpyxl.load_workbook(aq_file, data_only=True)
ws2 = wb2["Sheet1"]
headers = [ws2.cell(1, c).value for c in range(1, ws2.max_column+1)]
year_groups = []
for i, h in enumerate(headers):
    if h and isinstance(h,str) and re.search(r"[Tt]tl", h):
        m2 = re.search(r"(\d{2})$", h.strip())
        if m2: year_groups.append((2000+int(m2.group(1)), i))

print(f"  Years: {[y for y,_ in year_groups]}")
row_count = 0
for row in ws2.iter_rows(min_row=2, values_only=True):
    if not row[0]: continue
    m = re.match(r"(\d{2})\.(\d{2})\.(\d{2})\s+sale\s+(\d+)", str(row[0]), re.I)
    if m: row_count += 1

print(f"  Sale weeks parsed: {row_count}")
print("\nDry-run complete - no data written to Supabase")
